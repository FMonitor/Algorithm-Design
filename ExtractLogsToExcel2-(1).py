import openpyxl as xl
import openpyxl.styles as xls
import openpyxl.worksheet.worksheet as xlw
import os
import re
import datetime as dt

class Network:
    def __init__(self, method: str, classes: str, err_cnt: str,
                 ACC: str, BA: str, Kappa: str, AUC: str, F1: str,
                 PRE: str, REC: str, SPE: str, dateTime: str): 

        self.method: str = method
        self.classes = classes
        self.err_cnt = err_cnt
        self.ACC = ACC
        self.BA = BA
        self.Kappa = Kappa
        self.AUC = AUC
        self.F1 = F1
        self.PRE = PRE
        self.REC = REC
        self.SPE = SPE
        self.dateTime: str = dateTime


def read_log(log_file_path):
    err_cnt = ''
    ACC = ''
    BA = ''
    Kappa = ''
    AUC = ''
    F1 = ''
    PRE = ''
    REC = ''
    SPE = ''
    classes = ''
    dateTime = ''
    with open(log_file_path, 'r', encoding='UTF-8') as file:
        file_content = file.readlines()
        class_pattern = r"classes.*(\S+), (\S+)\]"
        for i, line in enumerate(file_content):
            match = re.search(class_pattern, line)
            if match:
                n1 = match.group(1)
                n2 = match.group(2)
                classes = n1+'-'+n2
            if re.search("-+AVERAGE SCORES", line):
                err_cnt = file_content[i+1].split(' ')[-1]
                err_cnt = err_cnt.split('\n')[0]
                pattern = r"-?\d+\.\d+±\d+\.\d+"
                match = re.findall(pattern, file_content[i+2])
                if match:
                    ACC = match[0]
                    BA = match[1]
                match = re.findall(pattern, file_content[i+3])
                if match:
                    Kappa = match[0]
                    AUC = match[1]
                match = re.findall(pattern, file_content[i+4])
                if match:
                    F1 = match[0]
                    PRE = match[1]
                match = re.findall(pattern, file_content[i+5])
                if match:
                    REC = match[0]
                    SPE = match[1]
                break
            if dateTime == '':
                dateTime = line[:19]

    return classes, err_cnt, ACC, BA, Kappa, AUC, F1, PRE, REC, SPE, dateTime


def read_folder(root_dir):
    networks = []
    for dirPath, dirNames, filenames in os.walk(root_dir):
        for filename in filenames:
            if filename.endswith('.log'):
                log_file_path = os.path.join(dirPath, filename)

                if os.path.isfile(log_file_path):
                    pattern = r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}).+\\([^\\]+).log"
                    match = re.search(pattern, str(log_file_path))
                    if match is not None:
                        date_time = match.group(1)  # get datetime： 2024-12-14_15-33-11
                        method = match.group(2)  # get method： ResNet-50
                    else:
                        pattern = r".+\\([^\\]+).log"
                        match = re.search(pattern, str(log_file_path))
                        assert match is not None, "Can not find a log file."
                        date_time = None
                        method = match.group(1)  # get method： ResNet-50

                    print(f"Reading file: {log_file_path}")

                    classes, err_cnt, ACC, BA, Kappa, AUC, F1, PRE, REC, SPE, dateTime = read_log(log_file_path)
                    assert classes != "", "`classes` not found!"
                    if date_time is None:
                        date_time = dateTime
                    network = Network(method, classes, err_cnt, ACC, BA, Kappa, AUC, F1, PRE, REC, SPE, date_time)
                    networks.append(network)
    return networks


def write_sheet(sheet: xlw.Worksheet, logs_dir: str):
    font = xls.Font(name="Times New Roman", size=12)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=12):
        for cell in row:
            cell.font = font

    networks = read_folder(logs_dir)

    # write header to Excel
    for i in range(1, 13):
        sheet.cell(row=1, column=i, value=["classes", "methods", "err_cnt", "ACC(%)", "BA(%)", "Kappa(%)", "AUC(%)", "F1(%)", "PRE(%)", "REC(%)", "SPE(%)", "DateTime"][i-1])

    # write data to Excel
    for i, network in enumerate(networks):
        row = i+2
        sheet.cell(row=row, column=1, value=network.classes)
        print(f"Writing {network.classes} {network.method} to Excel...")

        sheet.cell(row=row, column=2, value=network.method)
        cell = sheet.cell(row=row, column=3, value=float(network.err_cnt))
        cell.alignment = xls.Alignment(horizontal="left")
        sheet.cell(row=row, column=4, value=network.ACC)
        sheet.cell(row=row, column=5, value=network.BA)
        sheet.cell(row=row, column=6, value=network.Kappa)
        sheet.cell(row=row, column=7, value=network.AUC)
        sheet.cell(row=row, column=8, value=network.F1)
        sheet.cell(row=row, column=9, value=network.PRE)
        sheet.cell(row=row, column=10, value=network.REC)
        sheet.cell(row=row, column=11, value=network.SPE)
        sheet.cell(row=row, column=12, value=network.dateTime)

        # set column width
        for j in range(2, 12):
            sheet.column_dimensions[chr(64+j)].width = 14
        sheet.column_dimensions[chr(64+1)].width = 6
        sheet.column_dimensions[chr(64+3)].width = 8
        sheet.column_dimensions[chr(64+12)].width = 21
        # set alignment
        sheet.cell(row=row, column=3).alignment = xls.Alignment(horizontal="left", vertical="center")
    
    # delete the default sheet
    if "Sheet" in sheet.parent.sheetnames:
        sheet = sheet.parent["Sheet"]
        sheet.parent.remove(sheet)


def extract_logs_to_excel1(logs_dir: str, sheet_title: str, output_file_path: str):
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    write_sheet(sheet, logs_dir)
    
    save_workbook(workbook, output_file_path)


def extract_logs_to_excel2(logs_dir: str, output_file_path: str):
    workbook = xl.Workbook()
    for directory in os.listdir(logs_dir):
        print(f"Writing sheet {directory}...")
        sheet = workbook.create_sheet(directory)
        complete_path = os.path.join(logs_dir, directory)
        write_sheet(sheet, complete_path)

    save_workbook(workbook, output_file_path)


def save_workbook(workbook: xl.Workbook, output_file_path: str):
    # if output_file_path is not set, use the current time as the file name
    if output_file_path == "":
        strDatetime = dt.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        output_file_path = os.path.join(".", f"ExperimentalRecords_{strDatetime}.xlsx")

    workbook.save(output_file_path)
    print(f'Excel file saved to "{output_file_path}".')


if __name__ == "__main__":
    # ------------------------Arguments----------------------
    tRoot_dir = r".\logs\Network2_7 Room-Door"
    tSheet_title = ""
    tOutput_file_path = r".\documents\Ablation2_7 Room-Door.xlsx"
    # -------------------------------------------------------

    assert os.path.splitext(tOutput_file_path)[1] == ".xlsx", "The output file must be an xlsx file."
    if tSheet_title != "":
        extract_logs_to_excel1(tRoot_dir, tSheet_title, tOutput_file_path)
    else:
        extract_logs_to_excel2(tRoot_dir, tOutput_file_path)
